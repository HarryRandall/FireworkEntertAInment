#!/usr/bin/env python3
"""
Parse VDL shot sequences from a Hammer & Anvil XLSX and emit SQL for product_shots.

Handles the "1 Row" VDL format used by ICON Pyrotechnics:
  {N} Shot Cake (a) desc + (b) desc ..., 1 Row ({pan}{letter}{delay_ms}/…/CAK)

Each token encodes:
  - pan angle (integer degrees, may be negative)
  - effect letter (a-z, mapped to an effect_spec slug)
  - delay in ms to the NEXT shot (cumulative timing builds absolute fire times)

Skips:
  - Multi-row products (STR/STL/FNR/FNL layout — different format)
  - Fully encoded descriptions (TOP SECRET / {2KLUv/…} — no useful data)
  - Products that already have a parseable sequence but are listed as shells

Effect letter → effect spec slug mapping is derived from the (a) desc, (b) desc
sub-descriptions in the VDL header. Letters not found in the description fall
back to strobe-white.

Usage:
  python scripts/parse_vdl_shots.py "/path/to/Hammer & Anvil.xlsx"
  python scripts/parse_vdl_shots.py "/path/to/Hammer & Anvil.xlsx" --part HAC502407
  python scripts/parse_vdl_shots.py "/path/to/Hammer & Anvil.xlsx" --dry-run

Output is SQL printed to stdout — pipe to psql or a .sql file.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

DEFAULT_XLSX = Path("/Users/liammaloney/Downloads/Hammer & Anvil.xlsx")

# effect_spec slug → UUID in the live database
EFFECT_SPEC_IDS: dict[str, str] = {
    "fib-red":      "99ec1c08-26e2-40e7-a1c2-83374abd90f1",
    "fib-blue":     "8600cf7a-c308-499e-99c9-ccddbd10bd6d",
    "fib-green":    "86b306ce-9247-4eb8-8b46-2242c464b4c0",
    "fib-gold":     "50e8082f-aa61-4a35-9ef0-ab4cedd6e3a3",
    "strobe-mixed": "1b53d3ad-9564-4e23-a1c5-5b1263b51a29",
    "strobe-white": "6a43aa26-8167-4780-986e-9fda3069c9cb",
    "wave-purple":  "4427bde4-357d-4440-add5-7d699d74545c",
    "wave-rainbow": "f3847a97-787d-480a-9a20-27d736a1b724",
}


def desc_to_slug(desc: str | None) -> str:
    """Map a VDL sub-description string to an effect_spec slug."""
    if not desc or "{2KLUv/" in desc:
        return "strobe-white"
    d = desc.lower()
    has_red    = bool(re.search(r"\b(red|orange)\b", d))
    has_blue   = bool(re.search(r"\b(blue|indigo)\b", d))
    has_green  = bool(re.search(r"\bgreen\b", d))
    has_gold   = bool(re.search(r"\b(gold|lemon|yellow)\b", d))
    has_purple = bool(re.search(r"\b(purple|magenta)\b", d))
    n = sum([has_red, has_blue, has_green, has_gold, has_purple])
    if n >= 3:     return "wave-rainbow"
    if n == 2:     return "strobe-mixed"
    if has_purple: return "wave-purple"
    if has_gold:   return "fib-gold"
    if has_red:    return "fib-red"
    if has_blue:   return "fib-blue"
    if has_green:  return "fib-green"
    if re.search(r"\b(silver|white|brocade|willow|crackling)\b", d):
        return "strobe-white"
    return "strobe-mixed"


def build_letter_map(vdl: str) -> dict[str, str]:
    """
    Extract the (a) desc + (b) desc ... header from the VDL string and return
    a mapping of letter → effect_spec_slug.
    """
    letter_map: dict[str, str] = {}
    for m in re.finditer(r"\(([a-z])\)\s*([^+(,]+)", vdl):
        letter = m.group(1)
        description = m.group(2).strip()
        letter_map[letter] = desc_to_slug(description)
    return letter_map


def parse_1row_sequence(vdl: str) -> list[tuple[float, int, str]] | None:
    """
    Parse a '1 Row (…/CAK)' sequence from a VDL string.

    Returns a list of (time_seconds, pan_degrees, effect_spec_slug) tuples,
    or None if the VDL doesn't contain a parseable 1-row sequence.
    """
    m = re.search(r"1 Row \((.+?/CAK)\)", vdl, re.DOTALL)
    if not m:
        return None

    letter_map = build_letter_map(vdl)
    tokens = m.group(1).split("/")
    shots: list[tuple[float, int, str]] = []
    t_ms = 0

    for tok in tokens:
        tok = tok.strip()
        if not tok or tok == "CAK":
            break
        p = re.match(r"^(-?\d+)?([a-z])(\d+)?$", tok)
        if not p:
            continue
        pan    = int(p.group(1)) if p.group(1) else 0
        letter = p.group(2)
        delay  = int(p.group(3)) if p.group(3) else 0
        slug   = letter_map.get(letter, "strobe-white")
        shots.append((round(t_ms / 1000, 3), pan, slug))
        t_ms += delay

    return shots if shots else None


def load_rows(path: Path) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    out = []
    for raw in rows_iter:
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        row = dict(zip(headers, raw, strict=False))
        part = str(row.get("partNumber") or "").strip()
        if not part:
            continue
        out.append(row)
    return out


def caliber_for(row: dict) -> str:
    """Best-effort caliber from the spreadsheet row."""
    raw = str(row.get("caliber") or row.get("size") or "").strip()
    if not raw:
        pt = str(row.get("partType") or row.get("subtype") or "").lower()
        if "1.8" in pt or "1.8in" in pt:
            return "1.8in"
        return "30mm"
    raw = re.sub(r"\s+", "", raw)
    return raw


def emit_sql(part_number: str, caliber: str, shots: list[tuple[float, int, str]]) -> str:
    lines = [
        f"-- {part_number}",
        "DELETE FROM product_shots",
        f"  WHERE product_id = (SELECT id FROM products WHERE part_number = '{part_number}');",
        "",
    ]
    for idx, (t, pan, slug) in enumerate(shots, 1):
        spec_id = EFFECT_SPEC_IDS[slug]
        lines.append(
            f"INSERT INTO product_shots "
            f"(product_id, shot_index, caliber, time_offset_seconds, effect_spec_id, pan_degrees) VALUES "
            f"((SELECT id FROM products WHERE part_number = '{part_number}'), "
            f"{idx}, '{caliber}', {t}, '{spec_id}', {pan});"
        )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("xlsx", nargs="?", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--part", help="Only emit SQL for this part number")
    parser.add_argument("--dry-run", action="store_true", help="Print parse summary instead of SQL")
    args = parser.parse_args()

    path: Path = args.xlsx
    if not path or not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(1)

    rows = load_rows(path)
    skipped, parsed = [], []

    for row in rows:
        part = str(row.get("partNumber") or "").strip()
        if args.part and part != args.part:
            continue
        vdl = str(row.get("vdl") or row.get("VDL") or "").strip()
        if not vdl or "1 Row" not in vdl:
            skipped.append((part, "no 1-row sequence"))
            continue
        shots = parse_1row_sequence(vdl)
        if shots is None:
            skipped.append((part, "parse failed"))
            continue
        parsed.append((part, caliber_for(row), shots))

    if args.dry_run:
        print(f"Parsed {len(parsed)} products, skipped {len(skipped)}")
        for part, cal, shots in parsed:
            print(f"  {part}  caliber={cal}  shots={len(shots)}")
        if skipped:
            print("\nSkipped:")
            for part, reason in skipped:
                print(f"  {part}: {reason}")
        return

    for part, caliber, shots in parsed:
        print(emit_sql(part, caliber, shots))
        print()


if __name__ == "__main__":
    main()
